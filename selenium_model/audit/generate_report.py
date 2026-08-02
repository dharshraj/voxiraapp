"""Master Excel report builder — 20-sheet workbook.

Sheets:
 01 Executive Summary
 02 Selenium UI Tests
 03 Unit Tests
 04 API Validation Tests
 05 Form Validation Tests
 06 Performance Tests
 07 Load Test Results       (Locust)
 08 Vulnerability Tests
 09 Appium Mobile Tests
 10 E2E Flow Tests
 11 Functional Coverage
 12 Defect Report
 13 Security Observations
 14 Accessibility Findings
 15 Unused Files
 16 Dead Code
 17 Code Health Summary
 18 Recommendations
 19 Navigation Coverage
 20 Execution Summary

Run after pytest:
    python selenium_model/audit/generate_report.py
"""
from __future__ import annotations

import json
import sys
from datetime import datetime
from pathlib import Path

import pandas as pd

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
import config

try:
    import audit_data as ad
except ImportError:
    sys.path.insert(0, str(Path(__file__).resolve().parent))
    import audit_data as ad

OUTPUT = config.EXCEL_REPORT_PATH

# ── Colour palette ────────────────────────────────────────────────────────────
C = {
    "hdr_dark":  "1F3864",
    "hdr_blue":  "2E75B6",
    "hdr_green": "375623",
    "hdr_red":   "843C0C",
    "hdr_purple":"4B0082",
    "hdr_teal":  "005B5B",
    "hdr_grey":  "404040",
    "pass":   "C6EFCE",
    "fail":   "FFC7CE",
    "exc":    "FFEB9C",
    "skip":   "DDEBF7",
    "info":   "E2EFDA",
    "warn":   "FCE4D6",
    "alt":    "F2F2F2",
}

# ── Loaders ───────────────────────────────────────────────────────────────────

def _load(path: Path) -> list[dict]:
    if path.exists():
        try:
            return json.loads(path.read_text(encoding="utf-8"))
        except Exception:
            pass
    return []


def load_all_results() -> list[dict]:
    base = _load(config.RESULTS_JSON)
    seen = {r["test_id"] for r in base}
    for extra in (config.APPIUM_RESULTS_JSON, config.UNIT_RESULTS_JSON,
                  config.VULN_RESULTS_JSON):
        for r in _load(extra):
            if r["test_id"] not in seen:
                base.append(r)
                seen.add(r["test_id"])
    return base


def load_load_results() -> list[dict]:
    return _load(config.LOAD_RESULTS_JSON)


def _count_src_files() -> int:
    src = config.PROJECT_ROOT / "src"
    if not src.exists():
        return 0
    return sum(1 for p in src.rglob("*")
               if p.is_file() and p.suffix in {".ts", ".tsx", ".js", ".jsx"})


# ── Styling ───────────────────────────────────────────────────────────────────

def _style_ws(ws, df: pd.DataFrame,
              hdr_colour: str = "1F3864",
              row_colour_col: str | None = None) -> None:
    """Apply styling directly to an openpyxl Worksheet object."""
    try:
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return

    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    STATUS_FILL = {
        "Passed":       PatternFill("solid", fgColor=C["pass"]),
        "Failed":       PatternFill("solid", fgColor=C["fail"]),
        "Exception":    PatternFill("solid", fgColor=C["exc"]),
        "Skipped":      PatternFill("solid", fgColor=C["skip"]),
        "Open":         PatternFill("solid", fgColor=C["fail"]),
        "High":         PatternFill("solid", fgColor=C["fail"]),
        "Medium":       PatternFill("solid", fgColor=C["warn"]),
        "Low":          PatternFill("solid", fgColor=C["info"]),
        "Covered":      PatternFill("solid", fgColor=C["pass"]),
        "Not Covered":  PatternFill("solid", fgColor=C["warn"]),
    }
    ALT  = PatternFill("solid", fgColor=C["alt"])
    cols = list(df.columns)
    col_pos = cols.index(row_colour_col) + 1 if row_colour_col and row_colour_col in cols else None

    # Header row
    for ci, col in enumerate(cols, 1):
        cell = ws.cell(row=1, column=ci)
        cell.fill      = PatternFill("solid", fgColor=hdr_colour)
        cell.font      = Font(bold=True, color="FFFFFF", size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = thin
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

    # Data rows
    for ri in range(2, ws.max_row + 1):
        row_val = str(ws.cell(row=ri, column=col_pos).value or "") if col_pos else ""
        fill = STATUS_FILL.get(row_val, ALT if ri % 2 == 0 else None)
        for ci in range(1, len(cols) + 1):
            cell = ws.cell(row=ri, column=ci)
            if fill:
                cell.fill = fill
            cell.border    = thin
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        ws.row_dimensions[ri].height = 16

    # Column widths
    for ci, col in enumerate(cols, 1):
        try:
            w = max(df[col].astype(str).map(len).max(), len(str(col))) + 2
        except Exception:
            w = len(str(col)) + 2
        ws.column_dimensions[get_column_letter(ci)].width = min(w, 70)


def _style(writer, sheet_name: str, df: pd.DataFrame,
           hdr_colour: str = "1F3864",
           row_colour_col: str | None = None) -> None:
    """Legacy wrapper — delegates to _style_ws."""
    ws = writer.sheets.get(sheet_name)
    if ws:
        _style_ws(ws, df, hdr_colour=hdr_colour, row_colour_col=row_colour_col)


# ── Row builder ───────────────────────────────────────────────────────────────

def _rows(results: list[dict], *,
          module_filter: str | None = None,
          type_filter: str | None = None) -> list[dict]:
    out = []
    for i, r in enumerate(results, 1):
        if module_filter and r.get("module", "") != module_filter:
            continue
        if type_filter and r.get("test_type", "Selenium") != type_filter:
            continue
        out.append({
            "Test #":       i,
            "Test ID":      r.get("test_id", ""),
            "Module":       r.get("module", ""),
            "Type":         r.get("test_type", "Selenium"),
            "Scenario":     r.get("scenario", r.get("name", "")),
            "Expected":     r.get("expected", ""),
            "Actual":       r.get("actual", ""),
            "Status":       r.get("status", ""),
            "Duration (s)": r.get("duration_sec", ""),
            "Timestamp":    r.get("timestamp", ""),
            "Screenshot":   r.get("screenshot", ""),
        })
    if not out:
        out.append({k: ("No results — run the suite first" if k == "Scenario" else "")
                    for k in ["Test #", "Test ID", "Module", "Type", "Scenario",
                               "Expected", "Actual", "Status", "Duration (s)",
                               "Timestamp", "Screenshot"]})
    return out


def _df(results, **kw):
    return pd.DataFrame(_rows(results, **kw))


# ── Sheet 01: Executive Summary ───────────────────────────────────────────────

def s01_summary(results: list[dict], load: list[dict]) -> pd.DataFrame:
    total   = len(results)
    passed  = sum(1 for r in results if r["status"] == "Passed")
    failed  = sum(1 for r in results if r["status"] == "Failed")
    exc     = sum(1 for r in results if r["status"] == "Exception")
    skipped = sum(1 for r in results if r["status"] == "Skipped")
    rate    = f"{round(100*passed/total, 1)}%" if total else "0%"

    by_type: dict[str, dict] = {}
    for r in results:
        t = r.get("test_type", "Selenium")
        by_type.setdefault(t, {"p": 0, "t": 0})
        by_type[t]["t"] += 1
        if r["status"] == "Passed":
            by_type[t]["p"] += 1

    load_req  = sum(e.get("num_requests", 0) for e in load)
    load_fail = sum(e.get("num_failures", 0) for e in load)
    load_p95  = max((e.get("p95_ms", 0) for e in load), default=0)

    rows = [
        ("Report Generated",    datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("Project",             ad.PROJECT_NAME),
        ("Project Type",        ad.PROJECT_TYPE),
        ("Source Files (src/)", _count_src_files()),
        ("Screens / Pages",     len(ad.FUNCTIONALITY_MAP)),
        ("─── TEST RESULTS ───", ""),
        ("Total Tests Executed", total),
        ("Passed",              passed),
        ("Failed",              failed),
        ("Exceptions",          exc),
        ("Skipped",             skipped),
        ("Overall Pass Rate",   rate),
        ("─── BY TEST TYPE ───", ""),
    ]
    for ttype in sorted(by_type):
        d = by_type[ttype]
        rows.append((f"  {ttype}",
                     f"{d['p']}/{d['t']} ({round(100*d['p']/max(d['t'],1),1)}%)"))
    rows += [
        ("─── LOAD TEST ───",    ""),
        ("Total HTTP Requests",  load_req),
        ("Failed Requests",      load_fail),
        ("Failure Rate",         f"{round(100*load_fail/max(load_req,1),2)}%"),
        ("Peak p95 (ms)",        load_p95),
        ("─── CODE AUDIT ───",   ""),
        ("Confirmed Defects",    len(ad.LIVE_CONFIRMED_DEFECTS)),
        ("Unused Files",         len(ad.UNUSED_FILES)),
        ("Dead Code Instances",  len(ad.DEAD_CODE)),
        ("Security Findings",    len(ad.SECURITY_OBSERVATIONS)),
        ("─── REPORTS ───",      ""),
        ("Real-time Workbook",   str(config.ROOT_DIR / "REALTIME_TEST_PROGRESS.xlsx")),
        ("Master Report",        str(OUTPUT)),
    ]
    return pd.DataFrame(rows, columns=["Metric", "Value"])


# ── Sheets 02-10: Test result tables ─────────────────────────────────────────

def s02_selenium(results):
    rows = _rows(results)
    # keep only Selenium-type tests (exclude pure unit/API/appium)
    keep = [r for r in results
            if r.get("test_type", "Selenium") == "Selenium"
            and r.get("module") not in ("Unit", "Appium")]
    return pd.DataFrame(_rows(keep) if keep else _rows([], type_filter="__none__"))


def s03_unit(results):
    keep = [r for r in results if r.get("test_type") == "Unit"
            or r.get("module") == "Unit"]
    return pd.DataFrame(_rows(keep) if keep else _rows([], type_filter="__none__"))


def s04_api(results):
    keep = [r for r in results if r.get("test_type") == "API"
            or r.get("module") == "API Validation"]
    return pd.DataFrame(_rows(keep) if keep else _rows([], type_filter="__none__"))


def s05_validation(results):
    keep = [r for r in results if r.get("module") == "Validation"]
    return pd.DataFrame(_rows(keep) if keep else _rows([], type_filter="__none__"))


def s06_performance(results):
    keep = [r for r in results if r.get("module") == "Performance"]
    return pd.DataFrame(_rows(keep) if keep else _rows([], type_filter="__none__"))


def s07_load(load: list[dict]) -> pd.DataFrame:
    if not load:
        return pd.DataFrame([{
            "Endpoint": "No load results — run: python -m locust -f locust_load.py --headless",
            "Method": "", "Requests": 0, "Failures": 0, "Fail %": 0,
            "Avg ms": 0, "Min ms": 0, "Max ms": 0,
            "p50 ms": 0, "p90 ms": 0, "p95 ms": 0, "p99 ms": 0, "RPS": 0,
        }])
    return pd.DataFrame([{
        "Endpoint":  e.get("name", ""),
        "Method":    e.get("method", ""),
        "Requests":  e.get("num_requests", 0),
        "Failures":  e.get("num_failures", 0),
        "Fail %":    e.get("failure_rate_pct", 0),
        "Avg ms":    e.get("avg_response_ms", 0),
        "Min ms":    e.get("min_response_ms", 0),
        "Max ms":    e.get("max_response_ms", 0),
        "p50 ms":    e.get("p50_ms", 0),
        "p90 ms":    e.get("p90_ms", 0),
        "p95 ms":    e.get("p95_ms", 0),
        "p99 ms":    e.get("p99_ms", 0),
        "RPS":       e.get("rps", 0),
    } for e in load])


def s08_vulnerability(results):
    keep = [r for r in results if r.get("module") in ("Vulnerability", "Security Scan")
            or r.get("test_type") == "Security Scan"]
    return pd.DataFrame(_rows(keep) if keep else _rows([], type_filter="__none__"))


def s09_appium(results):
    keep = [r for r in results if r.get("test_type") == "Appium"
            or r.get("module") == "Appium"]
    return pd.DataFrame(_rows(keep) if keep else _rows([], type_filter="__none__"))


def s10_e2e(results):
    keep = [r for r in results if r.get("module") == "E2E Flow"]
    return pd.DataFrame(_rows(keep) if keep else _rows([], type_filter="__none__"))


# ── Sheet 11: Coverage ────────────────────────────────────────────────────────

def s11_coverage(results: list[dict]) -> pd.DataFrame:
    tested = " | ".join(r.get("scenario", "") for r in results).lower()
    rows = []
    for page, func, src in ad.FUNCTIONALITY_MAP:
        key = page.split(" ")[0].split("/")[0].lower().replace("screen", "")
        cov = "Covered" if key in tested else "Not Covered"
        note = ("Exercised by test suite" if cov == "Covered"
                else "Requires authenticated session or Appium device")
        rows.append({"Screen": page, "Functionality": func,
                     "Coverage Status": cov, "Source": src, "Notes": note})
    for name, reason in ad.NOT_APPLICABLE:
        rows.append({"Screen": "N/A", "Functionality": name,
                     "Coverage Status": "Not Applicable", "Source": "", "Notes": reason})
    return pd.DataFrame(rows)


# ── Sheet 12: Defects ─────────────────────────────────────────────────────────

def s12_defects(results: list[dict]) -> pd.DataFrame:
    rows = []
    n = 1
    for name, module, desc, sev in ad.LIVE_CONFIRMED_DEFECTS:
        rows.append({"Bug ID": f"BUG-{n:03d}", "Module": module.split(",")[0].strip(),
                     "Summary": name, "Evidence": desc[:300],
                     "Severity": sev, "Type": "Live Defect", "Status": "Open"})
        n += 1
    for r in results:
        if r["status"] not in ("Failed", "Exception"):
            continue
        rows.append({"Bug ID": f"BUG-{n:03d}", "Module": r.get("module", ""),
                     "Summary": r.get("scenario", ""),
                     "Evidence": f"pytest node: {r.get('test_id', '')}",
                     "Severity": "High" if r["status"] == "Exception" else "Medium",
                     "Type": "Test Failure", "Status": "Open"})
        n += 1
    for summary, module, steps, sev in [
        ("Broken barrel: SubscriptionScreen", "src/screens/settings/index",
         "Exports non-existent file", "High"),
        ("Broken barrel: LeaderboardScreen", "src/screens/profile/index",
         "Exports non-existent file", "High"),
        ("HelpScreen unreachable", "src/screens/settings/HelpScreen.tsx",
         "Built but not registered in navigator", "Medium"),
    ]:
        rows.append({"Bug ID": f"BUG-{n:03d}", "Module": module,
                     "Summary": summary, "Evidence": steps,
                     "Severity": sev, "Type": "Static Analysis", "Status": "Open"})
        n += 1
    if not rows:
        rows.append({k: "" for k in ["Bug ID", "Module", "Summary",
                                      "Evidence", "Severity", "Type", "Status"]})
    return pd.DataFrame(rows)


# ── Sheets 13-20 ──────────────────────────────────────────────────────────────

def s13_security():
    return pd.DataFrame([{"Area": a, "Observation": o, "Severity": s, "Recommendation": r}
                         for a, o, s, r in ad.SECURITY_OBSERVATIONS])


def s14_accessibility(results):
    rows = [{"Screen": "WelcomeScreen", "Issue": r.get("scenario", ""),
             "Severity": "Medium" if r["status"] != "Passed" else "Info",
             "Evidence": r.get("actual", ""),
             "Recommendation": "Add accessibilityLabel/accessibilityRole props"}
            for r in results if r.get("module") == "Accessibility"]
    rows.append({"Screen": "App-wide",
                 "Issue": "Zero aria-* attrs, zero semantic <button> tags — all divs with tabindex=0",
                 "Severity": "High",
                 "Evidence": "DOM probe confirmed",
                 "Recommendation": "Add accessibilityLabel/accessibilityRole to all Pressables"})
    return pd.DataFrame(rows)


def s15_unused():
    return pd.DataFrame([{"File": f, "Path": p, "Reason": r, "Severity": s}
                         for f, p, r, s in ad.UNUSED_FILES])


def s16_dead_code():
    return pd.DataFrame([{"File": f, "Symbol": sym, "Line": ln, "Recommendation": rec}
                         for f, sym, ln, rec in ad.DEAD_CODE])


def s17_code_health():
    rows = [{"Category": "Duplicate/Inconsistent", "Finding": f"{n}: {d}",
             "Severity": s, "Recommendation": "See Recommendations sheet"}
            for n, d, s in ad.DUPLICATE_OR_INCONSISTENT]
    rows.append({"Category": "Large File", "Finding": ad.LARGE_FILES_NOTE,
                 "Severity": "Medium", "Recommendation": "Split RootNavigator.tsx"})
    rows.append({"Category": "Dead Code", "Finding": f"{len(ad.DEAD_CODE)} instances",
                 "Severity": "Medium", "Recommendation": "See Dead Code sheet"})
    rows.append({"Category": "Unused Files", "Finding": f"{len(ad.UNUSED_FILES)} files",
                 "Severity": "Medium", "Recommendation": "See Unused Files sheet"})
    return pd.DataFrame(rows)


def s18_recommendations():
    return pd.DataFrame([{"Priority": p, "Recommendation": r, "Business Impact": b}
                         for p, r, b in ad.RECOMMENDATIONS])


def s19_navigation():
    data = [
        ("OnboardingStack", "Splash->Welcome->Feature1/2/3->GoalSelection->ExperienceLevel->Permissions->Register->Login->ForgotPassword", "Active — fully covered"),
        ("HomeStack", "Dashboard->DailyGoal->Notifications->Search->TopicDetail + Search deep-links", "Active — requires auth"),
        ("SpeechStack", "SpeechHome->Record->Analyzing->9 result screens->History->Dashboard->DailyChallenge->FillerWords->Pronunciation->PaceAndClarity + extras", "Active — requires auth"),
        ("ProfileStack", "Profile->EditProfile->ProgressOverview->Achievements->Settings->NotificationSettings->PrivacyPolicy->DeleteAccount->FAQ->Tutorial->WhatsNew", "Active — requires auth"),
        ("AchievementsTabStack", "AchievementsMain->Rewards->StreakCalendar->LevelUp", "Active — requires auth"),
        ("DailyGoalsStack", "DailyGoalMain->TopicDetail", "Active — requires auth"),
        ("GamificationStack", "Rewards->StreakCalendar->LevelUp", "DEAD — defined but never used in Tab.Navigator"),
        ("SupportStack", "FAQ->Tutorial->WhatsNew", "DEAD — defined but never used"),
    ]
    return pd.DataFrame([{"Stack": s, "Screens": sc, "Status": st} for s, sc, st in data])


def s20_exec_summary(results: list[dict]) -> pd.DataFrame:
    by_mod: dict[str, dict] = {}
    for r in results:
        m = r.get("module", "General")
        by_mod.setdefault(m, {"Passed": 0, "Failed": 0, "Exception": 0, "Skipped": 0, "Total": 0})
        by_mod[m]["Total"] += 1
        st = r.get("status", "Unknown")
        by_mod[m][st] = by_mod[m].get(st, 0) + 1
    rows = []
    for mod in sorted(by_mod):
        d = by_mod[mod]
        t = d["Total"]
        p = d.get("Passed", 0)
        rows.append({
            "Module":       mod,
            "Total":        t,
            "Passed":       p,
            "Failed":       d.get("Failed", 0),
            "Exception":    d.get("Exception", 0),
            "Skipped":      d.get("Skipped", 0),
            "Pass Rate %":  f"{round(100*p/t, 1)}%" if t else "0%",
        })
    if not rows:
        rows.append({"Module": "No results", "Total": 0, "Passed": 0,
                     "Failed": 0, "Exception": 0, "Skipped": 0, "Pass Rate %": "0%"})
    return pd.DataFrame(rows)


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    results = load_all_results()
    load    = load_load_results()
    defects = s12_defects(results)

    total  = len(results)
    passed = sum(1 for r in results if r.get("status") == "Passed")
    rate   = round(100 * passed / total, 1) if total else 0

    print(f"[report] {total} tests loaded | {passed} passed | {rate}% pass rate")
    print(f"[report] Writing -> {OUTPUT}")

    sheets = [
        ("01 Executive Summary",      s01_summary(results, load),  "1F3864", None),
        ("02 Selenium UI Tests",       s02_selenium(results),       "2E75B6", "Status"),
        ("03 Unit Tests",              s03_unit(results),           "375623", "Status"),
        ("04 API Validation Tests",    s04_api(results),            "2E75B6", "Status"),
        ("05 Form Validation Tests",   s05_validation(results),     "2E75B6", "Status"),
        ("06 Performance Tests",       s06_performance(results),    "404040", "Status"),
        ("07 Load Test Results",       s07_load(load),              "1F3864", None),
        ("08 Vulnerability Tests",     s08_vulnerability(results),  "843C0C", "Status"),
        ("09 Appium Mobile Tests",     s09_appium(results),         "375623", "Status"),
        ("10 E2E Flow Tests",          s10_e2e(results),            "2E75B6", "Status"),
        ("11 Functional Coverage",     s11_coverage(results),       "1F3864", "Coverage Status"),
        ("12 Defect Report",           defects,                     "843C0C", "Severity"),
        ("13 Security Observations",   s13_security(),              "843C0C", "Severity"),
        ("14 Accessibility Findings",  s14_accessibility(results),  "404040", "Severity"),
        ("15 Unused Files",            s15_unused(),                "404040", "Severity"),
        ("16 Dead Code",               s16_dead_code(),             "404040", None),
        ("17 Code Health",             s17_code_health(),           "404040", "Severity"),
        ("18 Recommendations",         s18_recommendations(),       "1F3864", "Priority"),
        ("19 Navigation Coverage",     s19_navigation(),            "1F3864", None),
        ("20 Execution Summary",       s20_exec_summary(results),   "1F3864", None),
    ]

    # Write data with xlsxwriter first
    with pd.ExcelWriter(str(OUTPUT), engine="xlsxwriter") as writer:
        for tab, df, _, _ in sheets:
            df.to_excel(writer, sheet_name=tab, index=False)

    # Re-open with openpyxl to apply styling
    try:
        import openpyxl
        from openpyxl import load_workbook

        wb = load_workbook(str(OUTPUT))
        for ws in wb.worksheets:
            ws.sheet_state = "visible"

        for tab, df, hdr, col in sheets:
            if tab in wb.sheetnames:
                ws = wb[tab]
                # Apply styling directly using openpyxl
                _style_ws(ws, df, hdr_colour=hdr, row_colour_col=col)

        wb.save(str(OUTPUT))
        print("[report] Styling applied successfully")
    except Exception as ex:
        print(f"[report] Styling note: {ex}")

    print(f"\n{'='*55}")
    print(f"  MASTER TEST AUDIT REPORT (20 sheets)")
    print(f"  {total} tests | {rate}% pass | {len(defects)} defects")
    print(f"  {len(load)} load entries")
    print(f"  Saved -> {OUTPUT}")
    print(f"{'='*55}\n")


if __name__ == "__main__":
    main()
