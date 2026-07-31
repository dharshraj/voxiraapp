"""Flat single-sheet Excel report — every individual test case as one row.

Unlike generate_report.py (20-sheet categorized audit workbook), this produces
ONE sheet containing every executed test case (unit + API + validation +
vulnerability + Selenium UI/navigation/performance/E2E + Appium, when present)
as a flat, sortable/filterable list — intended for reviewers who want to see
every one of the 1000+ individual test cases in a single view.

Run after pytest:
    python selenium_model/audit/generate_flat_report.py
"""
from __future__ import annotations

import json
import sys
from datetime import datetime
from pathlib import Path

import pandas as pd

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
import config

OUTPUT = config.ROOT_DIR / "ALL_TEST_CASES_FLAT_REPORT.xlsx"

C = {
    "hdr": "1F3864",
    "pass": "C6EFCE",
    "fail": "FFC7CE",
    "exc": "FFEB9C",
    "skip": "DDEBF7",
    "alt": "F2F2F2",
}


def _load(path: Path) -> list[dict]:
    if path.exists():
        try:
            return json.loads(path.read_text(encoding="utf-8"))
        except Exception:
            pass
    return []


def load_all_results() -> list[dict]:
    base = _load(config.RESULTS_JSON)
    seen = {r["test_id"] for r in base}
    for extra in (config.APPIUM_RESULTS_JSON, config.UNIT_RESULTS_JSON,
                  config.VULN_RESULTS_JSON):
        for r in _load(extra):
            if r["test_id"] not in seen:
                base.append(r)
                seen.add(r["test_id"])
    return base


def build_dataframe(results: list[dict]) -> pd.DataFrame:
    rows = []
    for i, r in enumerate(sorted(results, key=lambda x: x.get("test_id", "")), 1):
        rows.append({
            "#":            i,
            "Test ID":      r.get("test_id", ""),
            "Module":       r.get("module", "General"),
            "Type":         r.get("test_type", "Selenium"),
            "Scenario":     r.get("scenario", r.get("name", "")),
            "Expected":     r.get("expected", ""),
            "Actual":       r.get("actual", ""),
            "Status":       r.get("status", "Unknown"),
            "Duration (s)": r.get("duration_sec", ""),
            "Timestamp":    r.get("timestamp", ""),
        })
    if not rows:
        rows.append({k: ("No results — run the suite first" if k == "Scenario" else "")
                     for k in ["#", "Test ID", "Module", "Type", "Scenario",
                               "Expected", "Actual", "Status", "Duration (s)", "Timestamp"]})
    return pd.DataFrame(rows)


def _style(ws, df: pd.DataFrame) -> None:
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    thin = Border(left=Side(style="thin"), right=Side(style="thin"),
                  top=Side(style="thin"), bottom=Side(style="thin"))
    status_fill = {
        "Passed":    PatternFill("solid", fgColor=C["pass"]),
        "Failed":    PatternFill("solid", fgColor=C["fail"]),
        "Exception": PatternFill("solid", fgColor=C["exc"]),
        "Skipped":   PatternFill("solid", fgColor=C["skip"]),
    }
    alt = PatternFill("solid", fgColor=C["alt"])
    cols = list(df.columns)
    status_col = cols.index("Status") + 1

    for ci, _ in enumerate(cols, 1):
        cell = ws.cell(row=1, column=ci)
        cell.fill = PatternFill("solid", fgColor=C["hdr"])
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for ri in range(2, ws.max_row + 1):
        status_val = str(ws.cell(row=ri, column=status_col).value or "")
        fill = status_fill.get(status_val, alt if ri % 2 == 0 else None)
        for ci in range(1, len(cols) + 1):
            cell = ws.cell(row=ri, column=ci)
            if fill:
                cell.fill = fill
            cell.border = thin
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        ws.row_dimensions[ri].height = 16

    widths = {"#": 6, "Test ID": 55, "Module": 16, "Type": 12, "Scenario": 60,
              "Expected": 40, "Actual": 45, "Status": 10, "Duration (s)": 12, "Timestamp": 18}
    for ci, col in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(ci)].width = widths.get(col, 20)


def main():
    results = load_all_results()
    df = build_dataframe(results)

    total = len(results)
    passed = sum(1 for r in results if r.get("status") == "Passed")
    failed = sum(1 for r in results if r.get("status") == "Failed")
    exc = sum(1 for r in results if r.get("status") == "Exception")
    skipped = sum(1 for r in results if r.get("status") == "Skipped")
    rate = round(100 * passed / total, 1) if total else 0

    print(f"[flat-report] {total} test cases | {passed} passed | {failed} failed | "
          f"{exc} exception | {skipped} skipped | {rate}% pass rate")
    print(f"[flat-report] Writing -> {OUTPUT}")

    with pd.ExcelWriter(str(OUTPUT), engine="xlsxwriter") as writer:
        df.to_excel(writer, sheet_name="All Test Cases", index=False)

    from openpyxl import load_workbook
    wb = load_workbook(str(OUTPUT))
    ws = wb["All Test Cases"]
    _style(ws, df)
    wb.save(str(OUTPUT))

    print(f"[flat-report] Done — {total} rows written to sheet 'All Test Cases'")
    print(f"[flat-report] Generated at {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")


if __name__ == "__main__":
    main()
