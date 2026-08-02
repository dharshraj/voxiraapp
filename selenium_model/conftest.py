"""Master pytest configuration for VoxiraApp test framework.

Provides fixtures for:
  - Selenium WebDriver (chrome/firefox/edge, headed/headless, multi-viewport)
  - Appium mobile driver (Android / iOS)
  - Real-time Excel workbook that is updated after EVERY test (live progress)
  - JSON result sink used by audit/generate_report.py
  - Shared meta dict populated by each test function
  - Screenshot capture on pass/fail/exception
  - Browser console log dumping
"""

from __future__ import annotations

import json
import os
import sys
import time
import threading
from datetime import datetime
from pathlib import Path
from typing import Any, Generator

import pytest

sys.path.insert(0, str(Path(__file__).resolve().parent))

import config
from utils import screenshot as screenshot_util

# ---------------------------------------------------------------------------
# Real-time Excel workbook (openpyxl — written after every single test)
# ---------------------------------------------------------------------------
try:
    import openpyxl
    from openpyxl.styles import (
        PatternFill, Font, Alignment, Border, Side
    )
    _EXCEL_AVAILABLE = True
except ImportError:
    _EXCEL_AVAILABLE = False

_REALTIME_XLSX = config.ROOT_DIR / "REALTIME_TEST_PROGRESS.xlsx"
_EXCEL_LOCK = threading.Lock()

# Colour palette ─────────────────────────────────────────────────────────────
_FILL = {
    "header":    PatternFill("solid", fgColor="1F3864") if _EXCEL_AVAILABLE else None,
    "passed":    PatternFill("solid", fgColor="C6EFCE") if _EXCEL_AVAILABLE else None,
    "failed":    PatternFill("solid", fgColor="FFC7CE") if _EXCEL_AVAILABLE else None,
    "exception": PatternFill("solid", fgColor="FFEB9C") if _EXCEL_AVAILABLE else None,
    "skipped":   PatternFill("solid", fgColor="DDEBF7") if _EXCEL_AVAILABLE else None,
    "running":   PatternFill("solid", fgColor="FCE4D6") if _EXCEL_AVAILABLE else None,
    "alt_row":   PatternFill("solid", fgColor="F2F2F2") if _EXCEL_AVAILABLE else None,
}
_THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
) if _EXCEL_AVAILABLE else None

_XLSX_HEADERS = [
    "Test #", "Test ID", "Module", "Type", "Scenario",
    "Expected", "Actual", "Status", "Duration (s)",
    "Timestamp", "Screenshot",
]

def _init_realtime_wb() -> "openpyxl.Workbook | None":
    """Create a fresh real-time workbook with the header row."""
    if not _EXCEL_AVAILABLE:
        return None
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Live Test Results"
    ws.freeze_panes = "A2"

    for col_idx, header in enumerate(_XLSX_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill   = _FILL["header"]
        cell.font   = Font(bold=True, color="FFFFFF", size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _THIN_BORDER

    col_widths = [7, 40, 18, 14, 60, 40, 50, 10, 12, 18, 40]
    for col_idx, width in enumerate(col_widths, start=1):
        ws.column_dimensions[
            openpyxl.utils.get_column_letter(col_idx)
        ].width = width
    ws.row_dimensions[1].height = 24

    # Summary mini-table on a second sheet
    ws2 = wb.create_sheet("Summary")
    ws2["A1"], ws2["B1"] = "Metric", "Value"
    for c in ("A1", "B1"):
        ws2[c].fill = _FILL["header"]
        ws2[c].font = Font(bold=True, color="FFFFFF")
    for r, label in enumerate([
        "Total Tests", "Passed", "Failed", "Exception", "Skipped", "Pass Rate %",
        "Last Updated",
    ], start=2):
        ws2.cell(row=r, column=1, value=label)
    ws2.column_dimensions["A"].width = 20
    ws2.column_dimensions["B"].width = 20

    try:
        wb.save(str(_REALTIME_XLSX))
    except PermissionError:
        pass  # File locked by Excel — workbook stays in memory; flush will retry later
    return wb


def _status_fill(status: str):
    return _FILL.get(status.lower(), _FILL["running"])


_WB_CACHE: dict[str, Any] = {"wb": None, "dirty_count": 0}
_FLUSH_EVERY = 25  # write to disk every N results instead of on every single test


def _flush_realtime_wb(force: bool = False) -> None:
    """Save the in-memory workbook to disk. Cheap no-op if nothing pending.

    If the primary xlsx path is locked (e.g. open in Excel), silently write to
    a timestamped fallback file so the test run is never blocked.
    """
    wb = _WB_CACHE.get("wb")
    if wb is None or (not force and _WB_CACHE["dirty_count"] == 0):
        return
    try:
        wb.save(str(_REALTIME_XLSX))
        _WB_CACHE["dirty_count"] = 0
    except PermissionError:
        # Primary file is locked — write to a timestamped fallback silently
        fallback = _REALTIME_XLSX.with_name(
            f"REALTIME_TEST_PROGRESS_{int(time.time())}.xlsx"
        )
        try:
            wb.save(str(fallback))
            _WB_CACHE["dirty_count"] = 0
        except Exception:
            pass  # Never let Excel I/O crash the test run
    except Exception as exc:
        print(f"[conftest] real-time Excel flush failed: {exc}")


def _append_result_to_realtime_wb(result: dict[str, Any]) -> None:
    """Thread-safe append of one test result row to an in-memory workbook,
    flushed to disk periodically (not on every single test) — reloading and
    re-saving a growing .xlsx on every test does not scale past a few hundred
    tests, so the workbook is kept in memory for the whole session instead."""
    if not _EXCEL_AVAILABLE:
        return
    with _EXCEL_LOCK:
        try:
            wb = _WB_CACHE.get("wb")
            if wb is None:
                wb = _init_realtime_wb()
                _WB_CACHE["wb"] = wb
            if wb is None:
                return

            ws = wb["Live Test Results"]
            row = ws.max_row + 1
            fill = _status_fill(result.get("status", ""))
            is_alt = (row % 2 == 0)
            row_fill = fill if result["status"] != "Passed" else (
                _FILL["alt_row"] if is_alt else None
            )

            values = [
                row - 1,
                result.get("test_id", ""),
                result.get("module", ""),
                result.get("test_type", "Selenium"),
                result.get("scenario", ""),
                result.get("expected", ""),
                result.get("actual", ""),
                result.get("status", ""),
                result.get("duration_sec", ""),
                result.get("timestamp", ""),
                result.get("screenshot", ""),
            ]
            for col_idx, val in enumerate(values, start=1):
                cell = ws.cell(row=row, column=col_idx, value=val)
                if row_fill:
                    cell.fill = row_fill
                cell.border  = _THIN_BORDER
                cell.alignment = Alignment(
                    vertical="center", wrap_text=(col_idx in (5, 6, 7, 11))
                )
                if result.get("status") == "Failed":
                    cell.font = Font(color="9C0006")
                elif result.get("status") == "Exception":
                    cell.font = Font(color="9C6500")
            ws.row_dimensions[row].height = 18

            # Refresh summary sheet
            ws2 = wb["Summary"]
            all_rows = list(ws.iter_rows(min_row=2, values_only=True))
            total     = len(all_rows)
            passed    = sum(1 for r in all_rows if r[7] == "Passed")
            failed    = sum(1 for r in all_rows if r[7] == "Failed")
            exception = sum(1 for r in all_rows if r[7] == "Exception")
            skipped   = sum(1 for r in all_rows if r[7] == "Skipped")
            pass_rate = round(100 * passed / total, 1) if total else 0.0
            for r_idx, val in enumerate([
                total, passed, failed, exception, skipped, f"{pass_rate}%",
                datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            ], start=2):
                ws2.cell(row=r_idx, column=2, value=val)

            _WB_CACHE["dirty_count"] += 1
            if _WB_CACHE["dirty_count"] >= _FLUSH_EVERY:
                _flush_realtime_wb(force=True)
        except Exception as exc:
            # Never let Excel I/O crash the test run
            print(f"[conftest] real-time Excel write failed: {exc}")


# ---------------------------------------------------------------------------
# JSON result sink (consumed by audit/generate_report.py)
# ---------------------------------------------------------------------------
_results: list[dict[str, Any]] = []


# ---------------------------------------------------------------------------
# Session-scoped ChromeDriver path
# ---------------------------------------------------------------------------
@pytest.fixture(scope="session")
def chromedriver_path():
    from webdriver_manager.chrome import ChromeDriverManager
    return ChromeDriverManager().install()


# ---------------------------------------------------------------------------
# Selenium WebDriver fixture
# ---------------------------------------------------------------------------
@pytest.fixture
def driver(chromedriver_path):
    from selenium import webdriver as wd
    from selenium.webdriver.chrome.options import Options as ChromeOptions
    from selenium.webdriver.chrome.service import Service as ChromeService
    from selenium.webdriver.firefox.options import Options as FFOptions
    from selenium.webdriver.edge.options import Options as EdgeOptions

    browser = config.BROWSER.lower()

    if browser == "firefox":
        opts = FFOptions()
        if config.HEADLESS:
            opts.add_argument("--headless")
        drv = wd.Firefox(options=opts)
    elif browser == "edge":
        opts = EdgeOptions()
        if config.HEADLESS:
            opts.add_argument("--headless=new")
        drv = wd.Edge(options=opts)
    else:  # default: chrome
        opts = ChromeOptions()
        if config.HEADLESS:
            opts.add_argument("--headless=new")
        opts.add_argument(f"--window-size={config.WINDOW_WIDTH},{config.WINDOW_HEIGHT}")
        opts.add_argument("--disable-gpu")
        opts.add_argument("--no-sandbox")
        opts.add_argument("--disable-dev-shm-usage")
        opts.add_argument("--disable-extensions")
        opts.add_argument("--disable-popup-blocking")
        opts.set_capability("goog:loggingPrefs", {"browser": "ALL", "performance": "ALL"})
        drv = wd.Chrome(
            service=ChromeService(chromedriver_path), options=opts
        )

    drv.set_window_size(config.WINDOW_WIDTH, config.WINDOW_HEIGHT)
    drv.implicitly_wait(config.IMPLICIT_WAIT)
    drv.set_page_load_timeout(config.PAGE_LOAD_TIMEOUT)
    drv.get(config.BASE_URL)

    yield drv

    # Dump browser console logs
    try:
        logs = drv.get_log("browser")
        if logs:
            log_path = config.LOGS_DIR / f"console_{int(time.time()*1000)}.log"
            with open(log_path, "w", encoding="utf-8") as f:
                for entry in logs:
                    f.write(f"{entry.get('level')} {entry.get('message')}\n")
    except Exception:
        pass

    drv.quit()


# ---------------------------------------------------------------------------
# Responsive driver — takes a viewport preset name
# ---------------------------------------------------------------------------
@pytest.fixture
def responsive_driver(chromedriver_path):
    """Factory fixture: call responsive_driver('mobile_sm') to get a
    driver at that viewport size. Yields (driver, width, height)."""
    drivers_created: list = []

    def _make(preset_name: str):
        from selenium import webdriver as wd
        from selenium.webdriver.chrome.options import Options as ChromeOptions
        from selenium.webdriver.chrome.service import Service as ChromeService
        w, h = config.VIEWPORT_PRESETS.get(preset_name, (1440, 1024))
        opts = ChromeOptions()
        if config.HEADLESS:
            opts.add_argument("--headless=new")
        opts.add_argument(f"--window-size={w},{h}")
        opts.add_argument("--disable-gpu")
        opts.add_argument("--no-sandbox")
        opts.add_argument("--disable-dev-shm-usage")
        opts.set_capability("goog:loggingPrefs", {"browser": "ALL"})
        drv = wd.Chrome(
            service=ChromeService(chromedriver_path), options=opts
        )
        drv.set_window_size(w, h)
        drv.implicitly_wait(config.IMPLICIT_WAIT)
        drv.set_page_load_timeout(config.PAGE_LOAD_TIMEOUT)
        drv.get(config.BASE_URL)
        drivers_created.append(drv)
        return drv, w, h

    yield _make

    for d in drivers_created:
        try:
            d.quit()
        except Exception:
            pass


# ---------------------------------------------------------------------------
# Appium driver fixture (skipped if server unreachable or caps not set)
# ---------------------------------------------------------------------------
@pytest.fixture
def appium_driver(request):
    """Yields an Appium WebDriver for the platform specified by the
    'appium_platform' pytest mark (android | ios). Skips gracefully if
    Appium server is not reachable."""
    import urllib.request
    import urllib.error

    platform_mark = request.node.get_closest_marker("appium_platform")
    platform = (platform_mark.args[0] if platform_mark else "android").lower()

    # Connectivity check — skip entire test if Appium server is down
    try:
        urllib.request.urlopen(f"{config.APPIUM_SERVER_URL}/status", timeout=3)
    except Exception:
        pytest.skip(
            f"Appium server not reachable at {config.APPIUM_SERVER_URL} — "
            "start the Appium server and ensure the emulator/device is connected."
        )

    # Secondary check: verify the server is actually ready (not just port-open)
    # and that at least one device is listed as available. This prevents the
    # appium_wd.Remote() call below from hanging for 2+ minutes on a server
    # that accepted the TCP connection but has no device to create a session on.
    try:
        import json as _json
        with urllib.request.urlopen(
            f"{config.APPIUM_SERVER_URL}/status", timeout=5
        ) as _resp:
            _status_body = _json.loads(_resp.read())
        _ready = _status_body.get("value", {}).get("ready", False)
        if not _ready:
            pytest.skip(
                f"Appium server at {config.APPIUM_SERVER_URL} is online but "
                "reports ready=false — ensure a device/emulator is connected."
            )
    except Exception:
        pytest.skip(
            f"Appium server status check failed at {config.APPIUM_SERVER_URL}."
        )

    try:
        from appium import webdriver as appium_wd
        from appium.options.common import AppiumOptions
    except ImportError:
        pytest.skip("Appium-Python-Client not installed — run: pip install Appium-Python-Client")

    caps = (
        config.ANDROID_DESIRED_CAPS.copy()
        if platform == "android"
        else config.IOS_DESIRED_CAPS.copy()
    )

    # In Expo Go mode there is no APK path — skip the file check
    app_path = caps.get("appium:app", "")
    if app_path and not Path(app_path).exists():
        pytest.skip(
            f"App binary not found at {app_path} — build the app first "
            "(expo build / eas build) and set ANDROID_APP_PATH or IOS_APP_PATH."
        )

    options = AppiumOptions().load_capabilities(caps)

    # Wrap session creation in a thread with a 20-second hard timeout so a
    # connected-but-deviceless Appium server doesn't block the entire run.
    import concurrent.futures as _cf
    def _create_session():
        return appium_wd.Remote(config.APPIUM_SERVER_URL, options=options)

    with _cf.ThreadPoolExecutor(max_workers=1) as _pool:
        _future = _pool.submit(_create_session)
        try:
            drv = _future.result(timeout=60)
        except _cf.TimeoutError:
            pytest.skip(
                "Appium session creation timed out after 60 s — "
                "ensure a device/emulator is booted and connected."
            )
        except Exception as _sess_err:
            pytest.skip(
                f"Appium session could not be created: {_sess_err}. "
                "Ensure the device/emulator is booted and Expo Go is installed."
            )
    drv.implicitly_wait(config.APPIUM_IMPLICIT_WAIT)

    if platform == "android":
        # Some devices ship with rotation lock enabled, which makes Appium's
        # simulated `.orientation = "LANDSCAPE"` calls fail outright. Force
        # auto-rotate on so orientation-change tests can actually take effect.
        import subprocess
        try:
            subprocess.run(
                ["adb", "-s", config.ANDROID_DEVICE_NAME, "shell",
                 "settings", "put", "system", "accelerometer_rotation", "1"],
                capture_output=True, timeout=10,
            )
        except Exception:
            pass

    if platform == "android" and config.EXPO_GO_MODE:
        # With noReset=true, a new Appium *session* does not kill Expo Go's
        # existing JS process or clear its storage — it just brings the same
        # running experience to the foreground, wherever the PREVIOUS test's
        # navigation (or auth session!) left it. A plain terminate_app() only
        # kills the process, which is not enough: if a prior test logged in
        # (e.g. the authenticated tab-bar test), the persisted Supabase
        # session in AsyncStorage survives the kill, and the app auto-redirects
        # straight to the Dashboard on relaunch instead of the Welcome screen
        # every later test expects. `pm clear` wipes app storage too,
        # guaranteeing every test starts fully logged-out and cold.
        import subprocess as _subprocess
        try:
            _subprocess.run(
                ["adb", "-s", config.ANDROID_DEVICE_NAME, "shell",
                 "pm", "clear", "host.exp.exponent"],
                capture_output=True, timeout=15,
            )
        except Exception:
            pass
        time.sleep(1)

        # Expo Go's own launcher activity does not process a bundled deep-link
        # URL in its launch intent on current builds — it just opens Expo Go's
        # home screen. `mobile: deepLink` issues a plain, unforced VIEW intent
        # (no explicit component, no extra flags) which Android resolves to
        # Expo Go's ExperienceActivity correctly, matching what was confirmed
        # to work via `adb shell am start -a VIEW -d exp://...`.
        drv.execute_script("mobile: deepLink", {
            "url": config.EXPO_LAN_URL,
            "package": "host.exp.exponent",
        })
        # Wait for Expo Go to bundle and render — 2 s is not enough on cold start.
        # Poll for the WelcomeScreen heading text (up to 30 s) so each test
        # begins with the app actually loaded rather than the Expo splash/loader.
        from appium.webdriver.common.appiumby import AppiumBy
        from selenium.webdriver.support.ui import WebDriverWait
        from selenium.webdriver.support import expected_conditions as EC
        try:
            WebDriverWait(drv, 30).until(
                EC.presence_of_element_located(
                    (AppiumBy.XPATH,
                     "//*[@text='Get Started Free' or @text='MASTER' or "
                     "@text='Sign In' or @content-desc='Get Started Free']")
                )
            )
        except Exception:
            time.sleep(8)  # fallback flat wait if XPath probe fails

    if platform == "android" and not config.EXPO_GO_MODE:
        # Real standalone build: same test-isolation problem as Expo Go mode
        # (noReset=true means a prior test's login session can persist across
        # runs), just against the app's own package instead of Expo Go's.
        import subprocess as _subprocess
        try:
            _subprocess.run(
                ["adb", "-s", config.ANDROID_DEVICE_NAME, "shell",
                 "pm", "clear", config.ANDROID_PACKAGE_NAME],
                capture_output=True, timeout=15,
            )
        except Exception:
            pass
        time.sleep(1)
        try:
            drv.activate_app(config.ANDROID_PACKAGE_NAME)
        except Exception:
            pass

        from appium.webdriver.common.appiumby import AppiumBy
        from selenium.webdriver.support.ui import WebDriverWait
        from selenium.webdriver.support import expected_conditions as EC
        try:
            WebDriverWait(drv, 30).until(
                EC.presence_of_element_located(
                    (AppiumBy.XPATH,
                     "//*[@text='Get Started Free' or @text='MASTER' or "
                     "@text='Sign In' or @content-desc='Get Started Free']")
                )
            )
        except Exception:
            time.sleep(8)

    yield drv

    # Gracefully quit the Appium session. If the session was already terminated
    # during the test (e.g. the device killed the process, or the test called
    # terminate_app()), drv.quit() throws InvalidSessionIdException which shows
    # up as a spurious ERROR in the pytest report.  Swallow it silently.
    try:
        drv.quit()
    except Exception:
        pass


# ---------------------------------------------------------------------------
# Shared meta fixture — every test populates this
# ---------------------------------------------------------------------------
@pytest.fixture
def meta():
    return {
        "module":    "General",
        "test_type": "Selenium",
        "scenario":  "",
        "expected":  "",
        "actual":    "",
    }


# ---------------------------------------------------------------------------
# pytest hooks: capture result, write Excel row, write JSON
# ---------------------------------------------------------------------------
@pytest.hookimpl(tryfirst=True, hookwrapper=True)
def pytest_runtest_makereport(item, call):
    outcome = yield
    report  = outcome.get_result()

    if report.when != "call":
        return

    duration  = round(report.duration, 3)
    meta_dict = item.funcargs.get("meta", {}) if hasattr(item, "funcargs") else {}
    drv       = item.funcargs.get("driver") if hasattr(item, "funcargs") else None

    if report.passed:
        status = "Passed"
    elif report.failed:
        status = "Failed" if (
            call.excinfo is not None and call.excinfo.errisinstance(AssertionError)
        ) else "Exception"
    elif report.skipped:
        status = "Skipped"
    else:
        status = "Unknown"

    screenshot_path = ""
    if drv is not None and status in ("Passed", "Failed", "Exception"):
        bucket = {"Passed": "passed", "Failed": "failed", "Exception": "exceptions"}[status]
        screenshot_path = screenshot_util.capture(drv, item.name, bucket)

    actual = meta_dict.get("actual", "")
    if not actual and report.failed and call.excinfo is not None:
        actual = str(call.excinfo.value)[:500]

    record = {
        "test_id":      item.nodeid,
        "name":         item.name,
        "module":       meta_dict.get("module", "General"),
        "test_type":    meta_dict.get("test_type", "Selenium"),
        "scenario":     meta_dict.get("scenario", item.name),
        "expected":     meta_dict.get("expected", ""),
        "actual":       actual or ("As expected" if status == "Passed" else ""),
        "status":       status,
        "duration_sec": duration,
        "timestamp":    datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "screenshot":   screenshot_path,
    }

    _results.append(record)

    # Real-time Excel update (non-blocking append)
    _append_result_to_realtime_wb(record)


def pytest_sessionfinish(session, exitstatus):
    """Persist all results to JSON for the Excel report builder."""
    _flush_realtime_wb(force=True)

    # Merge with any pre-existing results (Appium, unit, load run separately)
    existing: list[dict] = []
    if config.RESULTS_JSON.exists():
        try:
            with open(config.RESULTS_JSON, encoding="utf-8") as f:
                existing = json.load(f)
        except Exception:
            existing = []

    # De-duplicate by test_id; new run wins
    merged = {r["test_id"]: r for r in existing}
    merged.update({r["test_id"]: r for r in _results})

    with open(config.RESULTS_JSON, "w", encoding="utf-8") as f:
        json.dump(list(merged.values()), f, indent=2)

    total  = len(_results)
    passed = sum(1 for r in _results if r["status"] == "Passed")
    print(
        f"\n[conftest] Session complete — {total} tests, "
        f"{passed} passed. "
        f"Real-time workbook: {_REALTIME_XLSX}\n"
        f"JSON sink: {config.RESULTS_JSON}"
    )
